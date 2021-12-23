import pandas as pd
import openpyxl
import json


class Converter:
    @staticmethod
    def excel2json(excel_filepath, sheet_name, output_filename=None,
                   orient='records'):
        # load the excel file
        workbook = openpyxl.load_workbook(excel_filepath)
        # get the sheet
        worksheet = workbook[sheet_name]
        # get all the column names
        num_cols = worksheet.max_column
        columns = [worksheet.cell(1, i).value
                   for i in range(1, num_cols+1)]
        
        # orient type is index
        if orient == 'index':
            json_data = {}
            for i, values in enumerate(worksheet.values):
                if i == 0:
                    continue
                data = {}

                for j, col in enumerate(columns):
                    data[col] = values[j]
                json_data[str(i-1)] = data
            
            if output_filename is not None:
                with open(output_filename, "w") as f:
                    json.dump(json_data, f)

            return json_data
        
        elif orient == 'records':
            # list to store data
            json_data = []
            for i, values in enumerate(worksheet.values):
                if i == 0:
                    continue
                data = {}

                for j, col in enumerate(columns):
                    data[col] = values[j]
                json_data.append(data)

            if output_filename is not None:
                with open(output_filename, "w") as f:
                    json.dump(json_data, f)

            return json_data
        
        # if orient type is 'columns'
        elif orient == 'columns':
            json_data = {}
            
            for j, col in enumerate(columns):
                json_data[col] = {}
                
                for i, values in enumerate(worksheet.values):
                    if i == 0:
                        continue

                    json_data[col][str(i-1)] = values[j]

            if output_filename is not None:
                with open(output_filename, "w") as f:
                    json.dump(json_data, f)
                    
            return json_data
                